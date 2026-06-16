"""
StockSense 抽圖後端 v2
功能:接收 Excel,抽出內嵌圖片並對應「款號+色號」,上傳到 Supabase Storage,回傳結果。
修正:之前版本用 style_no 識別,導致同款不同色共用第一張圖。
   現在用 (style_no, color_code) 組合識別,每個色號獨立一張圖。
"""
import os, re, zipfile, tempfile, subprocess, shutil, uuid
from typing import List, Dict, Optional
from fastapi import FastAPI, UploadFile, File, HTTPException, Header, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, Response
import openpyxl
import httpx

app = FastAPI(title="StockSense Image Extractor")

@app.middleware("http")
async def add_cors_headers(request: Request, call_next):
    if request.method == "OPTIONS":
        response = Response(status_code=200)
        response.headers["Access-Control-Allow-Origin"] = "*"
        response.headers["Access-Control-Allow-Methods"] = "GET, POST, OPTIONS"
        response.headers["Access-Control-Allow-Headers"] = "*"
        response.headers["Access-Control-Max-Age"] = "86400"
        return response
    response = await call_next(request)
    response.headers["Access-Control-Allow-Origin"] = "*"
    response.headers["Access-Control-Allow-Methods"] = "GET, POST, OPTIONS"
    response.headers["Access-Control-Allow-Headers"] = "*"
    return response

SUPABASE_URL = os.environ.get("SUPABASE_URL", "")
SUPABASE_SERVICE_KEY = os.environ.get("SUPABASE_SERVICE_KEY", "")
SUPABASE_JWT_SECRET = os.environ.get("SUPABASE_JWT_SECRET", "")

BUCKET = "product-images"


async def verify_token(authorization: Optional[str] = Header(None)) -> Dict:
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="缺少 Authorization header")
    token = authorization.split(" ", 1)[1]
    async with httpx.AsyncClient(timeout=10) as client:
        r = await client.get(
            f"{SUPABASE_URL}/auth/v1/user",
            headers={"Authorization": f"Bearer {token}", "apikey": SUPABASE_SERVICE_KEY}
        )
    if r.status_code != 200:
        raise HTTPException(status_code=401, detail=f"Token 驗證失敗: {r.text[:100]}")
    return r.json()


def convert_xls_to_xlsx(input_path: str, output_dir: str) -> str:
    result = subprocess.run(
        ["libreoffice", "--headless", "--convert-to", "xlsx", "--outdir", output_dir, input_path],
        capture_output=True, text=True, timeout=120
    )
    if result.returncode != 0:
        raise RuntimeError(f"LibreOffice 轉檔失敗: {result.stderr}")
    base = os.path.splitext(os.path.basename(input_path))[0]
    out = os.path.join(output_dir, base + ".xlsx")
    if not os.path.exists(out):
        raise RuntimeError("LibreOffice 轉檔後找不到 .xlsx 檔")
    return out


def extract_color_code(cell_value) -> str:
    """從「顏色」欄字串取出色號(例如 'BIH01 大愛 心 p752...' -> 'BIH01')"""
    if cell_value is None:
        return ""
    s = str(cell_value).strip()
    m = re.match(r'^([A-Z]{2,4}\d{1,4}[A-Z0-9]*)', s)
    return m.group(1) if m else ""


def extract_images_with_styles(xlsx_path: str) -> List[Dict]:
    """
    從 xlsx 抽圖並對應「款號+色號」。
    回傳: [{style_no, color_code, image_bytes, image_ext}, ...]
    """
    results = []
    seen_keys = set()  # 改成 (style, color) tuple,每個色號獨立追蹤
    z = zipfile.ZipFile(xlsx_path)
    names = z.namelist()
    ws_files = sorted([n for n in names if re.match(r'xl/worksheets/sheet\d+\.xml$', n)])
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheetnames = wb.sheetnames

    for idx, wsfile in enumerate(ws_files):
        sn = sheetnames[idx] if idx < len(sheetnames) else "?"
        if 'A品' not in sn and 'A 品' not in sn:
            continue
        rels_path = f"xl/worksheets/_rels/{os.path.basename(wsfile)}.rels"
        if rels_path not in names:
            continue
        dm = re.search(r'drawing(\d+)\.xml', z.read(rels_path).decode('utf-8', 'ignore'))
        if not dm:
            continue
        draw = f"xl/drawings/drawing{dm.group(1)}.xml"
        if draw not in names:
            continue
        dcontent = z.read(draw).decode('utf-8', 'ignore')
        anchors = re.findall(
            r'<xdr:from>.*?<xdr:row>(\d+)</xdr:row>.*?r:embed="(rId\d+)"',
            dcontent, re.S
        )
        drels_path = f"xl/drawings/_rels/drawing{dm.group(1)}.xml.rels"
        rid2img = {}
        if drels_path in names:
            rid2img = {
                m[0]: m[1] for m in re.findall(
                    r'Id="(rId\d+)"[^>]*Target="\.\./media/([^"]+)"',
                    z.read(drels_path).decode('utf-8', 'ignore')
                )
            }
        ws = wb[sn]
        # 同時抓款號(第 4 欄) 和 色號(第 5 欄)
        row2style = {}
        row2color = {}
        for r in range(1, ws.max_row + 1):
            v = ws.cell(r, 4).value
            if v:
                raw = str(v).strip()
                m = re.match(r'([A-Z]{2,4}\d[A-Z0-9]*)', raw)
                if m:
                    row2style[r - 1] = m.group(1)
            cv = ws.cell(r, 5).value
            if cv:
                cc = extract_color_code(cv)
                if cc:
                    row2color[r - 1] = cc

        for row, rid in anchors:
            style = row2style.get(int(row))
            color = row2color.get(int(row), "")
            img = rid2img.get(rid)
            if not style or not img:
                continue
            key = (style, color)
            if key in seen_keys:
                continue
            ext = img.split('.')[-1].lower()
            if ext == 'jpeg':
                ext = 'jpg'
            img_data = z.read('xl/media/' + img)
            results.append({
                "style_no": style,
                "color_code": color,
                "part": "",
                "image_bytes": img_data,
                "image_ext": ext,
            })
            seen_keys.add(key)
    return results


async def upload_to_supabase_storage(image_bytes: bytes, filename: str) -> str:
    url = f"{SUPABASE_URL}/storage/v1/object/{BUCKET}/{filename}"
    ext = filename.split('.')[-1].lower()
    content_type = "image/jpeg" if ext in ("jpg","jpeg") else f"image/{ext}"
    async with httpx.AsyncClient(timeout=30) as client:
        r = await client.post(
            url, content=image_bytes,
            headers={
                "Authorization": f"Bearer {SUPABASE_SERVICE_KEY}",
                "apikey": SUPABASE_SERVICE_KEY,
                "Content-Type": content_type,
                "x-upsert": "true",
            },
        )
        if r.status_code not in (200, 201):
            raise RuntimeError(f"Storage 上傳失敗 ({r.status_code}): {r.text[:200]}")
    return f"{SUPABASE_URL}/storage/v1/object/public/{BUCKET}/{filename}"


@app.get("/")
def health():
    return {"ok": True, "service": "StockSense Image Extractor", "version": "v2_color_aware"}


@app.options("/extract")
def extract_preflight():
    return {"ok": True}


@app.post("/extract")
async def extract(
    file: UploadFile = File(...),
    authorization: Optional[str] = Header(None),
):
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="缺少 Authorization header")
    access_token = authorization.split(" ", 1)[1]
    await verify_token(authorization)

    tmpdir = tempfile.mkdtemp(prefix="ssx_")
    try:
        in_path = os.path.join(tmpdir, file.filename)
        content = await file.read()
        with open(in_path, "wb") as f:
            f.write(content)

        if file.filename.lower().endswith(".xls"):
            xlsx_path = convert_xls_to_xlsx(in_path, tmpdir)
        elif file.filename.lower().endswith(".xlsx"):
            xlsx_path = in_path
        else:
            raise HTTPException(status_code=400, detail="只接受 .xls 或 .xlsx 檔")

        images = extract_images_with_styles(xlsx_path)

        results = []
        for img in images:
            safe_style = re.sub(r'[^A-Za-z0-9_-]', '', img['style_no'])
            safe_color = re.sub(r'[^A-Za-z0-9_-]', '', img.get('color_code', ''))
            unique_name = f"{safe_style}_{safe_color}_{uuid.uuid4().hex[:6]}.{img['image_ext']}"
            try:
                url = await upload_to_supabase_storage(img["image_bytes"], unique_name)
                results.append({
                    "style_no": img["style_no"],
                    "color_code": img.get("color_code", ""),
                    "image_url": url,
                    "filename": unique_name,
                })
                print(f"✅ 上傳成功: {img['style_no']}/{img.get('color_code','')} -> {unique_name}")
            except Exception as e:
                err_msg = str(e)
                results.append({
                    "style_no": img["style_no"],
                    "color_code": img.get("color_code", ""),
                    "error": err_msg,
                })
                print(f"❌ 上傳失敗: {img['style_no']}/{img.get('color_code','')} -> {err_msg}")

        return {
            "ok": True,
            "extracted": len(images),
            "uploaded": len([r for r in results if "image_url" in r]),
            "results": results,
        }
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)
